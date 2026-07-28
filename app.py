import streamlit as st
import pandas as pd
from io import BytesIO, StringIO
from openpyxl.styles import PatternFill
import re
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows

# Custom CSS for better styling
st.markdown("""
    <style>
    /* Main container styling */
    .main {
        padding: 0rem 1rem;
    }
    
    /* Title styling */
    .main-title {
        font-size: 2.5rem !important;
        font-weight: 700 !important;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        padding: 1rem 0;
    }
    
    /* Card styling */
    .card {
        background: white;
        padding: 1.5rem;
        border-radius: 10px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        margin: 1rem 0;
        border-left: 4px solid #667eea;
    }
    
    .card-success {
        border-left-color: #28a745;
    }
    
    .card-warning {
        border-left-color: #ffc107;
    }
    
    .card-danger {
        border-left-color: #dc3545;
    }
    
    .card-info {
        border-left-color: #17a2b8;
    }
    
    /* Metric container */
    .metric-container {
        background: white;
        padding: 1.2rem;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0, 0, 0, 0.05);
        text-align: center;
        transition: transform 0.2s;
        margin: 0.5rem 0;
    }
    
    .metric-container:hover {
        transform: translateY(-5px);
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.1);
    }
    
    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        color: #2c3e50;
    }
    
    .metric-label {
        font-size: 0.9rem;
        color: #7f8c8d;
        margin-top: 0.5rem;
    }
    
    .metric-icon {
        font-size: 2rem;
        margin-bottom: 0.5rem;
    }
    
    /* Status badges */
    .badge {
        display: inline-block;
        padding: 0.25rem 0.75rem;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: 600;
    }
    
    .badge-success {
        background: #d4edda;
        color: #155724;
    }
    
    .badge-danger {
        background: #f8d7da;
        color: #721c24;
    }
    
    .badge-warning {
        background: #fff3cd;
        color: #856404;
    }
    
    .badge-info {
        background: #d1ecf1;
        color: #0c5460;
    }
    
    /* Button styling */
    .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        padding: 0.6rem 1.5rem;
        border-radius: 8px;
        font-weight: 600;
        transition: all 0.3s;
        width: 100%;
    }
    
    .stButton > button:hover {
        transform: scale(1.02);
        box-shadow: 0 4px 8px rgba(102, 126, 234, 0.4);
    }
    
    .stButton > button:focus {
        box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.3);
    }
    
    /* Download button styling */
    .download-btn {
        background: linear-gradient(135deg, #28a745 0%, #20c997 100%);
        color: white;
        border: none;
        padding: 0.6rem 1.5rem;
        border-radius: 8px;
        font-weight: 600;
        transition: all 0.3s;
        width: 100%;
    }
    
    .download-btn:hover {
        transform: scale(1.02);
        box-shadow: 0 4px 8px rgba(40, 167, 69, 0.4);
    }
    
    /* File uploader styling */
    .upload-container {
        border: 2px dashed #667eea;
        border-radius: 10px;
        padding: 2rem;
        text-align: center;
        background: #f8f9fa;
        transition: all 0.3s;
    }
    
    .upload-container:hover {
        background: #e9ecef;
        border-color: #764ba2;
    }
    
    /* Divider styling */
    .custom-divider {
        height: 3px;
        background: linear-gradient(90deg, #667eea, #764ba2, #667eea);
        margin: 2rem 0;
        border-radius: 3px;
    }
    
    /* Expander styling */
    .streamlit-expanderHeader {
        font-weight: 600;
        color: #2c3e50;
    }
    
    /* Dataframe styling */
    .dataframe {
        border-radius: 10px;
        overflow: hidden;
    }
    
    /* Progress bar styling */
    .stProgress > div > div {
        background: linear-gradient(90deg, #667eea, #764ba2);
    }
    
    /* Footer styling */
    .footer {
        text-align: center;
        color: #7f8c8d;
        padding: 2rem 0;
        border-top: 1px solid #e9ecef;
        margin-top: 2rem;
    }
    
    /* Responsive adjustments */
    @media (max-width: 768px) {
        .metric-value {
            font-size: 1.5rem;
        }
        .main-title {
            font-size: 1.8rem !important;
        }
    }
    </style>
""", unsafe_allow_html=True)

st.set_page_config(
    page_title="Duplicate & Invalid Number Checker",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Initialize ALL session state variables
if 'processed' not in st.session_state:
    st.session_state.processed = False
if 'df' not in st.session_state:
    st.session_state.df = None
if 'output_data' not in st.session_state:
    st.session_state.output_data = None
if 'output_data_csv' not in st.session_state:
    st.session_state.output_data_csv = None
if 'clean_output_data' not in st.session_state:
    st.session_state.clean_output_data = None
if 'clean_output_data_csv' not in st.session_state:
    st.session_state.clean_output_data_csv = None
if 'file_uploaded' not in st.session_state:
    st.session_state.file_uploaded = False
if 'clean_df' not in st.session_state:
    st.session_state.clean_df = None
if 'clean_count' not in st.session_state:
    st.session_state.clean_count = 0
if 'duplicate_count' not in st.session_state:
    st.session_state.duplicate_count = 0
if 'duplicate_count_before' not in st.session_state:
    st.session_state.duplicate_count_before = 0
if 'invalid_count' not in st.session_state:
    st.session_state.invalid_count = 0
if 'valid_count' not in st.session_state:
    st.session_state.valid_count = 0
if 'column' not in st.session_state:
    st.session_state.column = None
if 'sheet_names' not in st.session_state:
    st.session_state.sheet_names = None
if 'selected_sheet' not in st.session_state:
    st.session_state.selected_sheet = None
if 'original_file' not in st.session_state:
    st.session_state.original_file = None

# Custom title with gradient
st.markdown('<h1 class="main-title">📱 Duplicate & Invalid Mobile Number Checker</h1>', unsafe_allow_html=True)

# Info card
st.markdown("""
<div class="card card-info">
    <h4>🚀 Quick Guide</h4>
    <ul style="margin-bottom: 0;">
        <li>Upload your Excel or CSV file containing mobile numbers</li>
        <li>For Excel files with multiple sheets, select the desired sheet</li>
        <li>Select the column with mobile numbers</li>
        <li>Configure validation settings</li>
        <li>Click <strong>Process File</strong> to analyze</li>
        <li>Download results in <strong>Excel</strong> or <strong>CSV</strong> format</li>
    </ul>
</div>
""", unsafe_allow_html=True)

uploaded_file = st.file_uploader(
    "📂 Upload Excel or CSV File",
    type=["xlsx", "csv"],
    help="Supported formats: .xlsx, .csv",
    key="file_uploader"
)

# Reset processed state when new file is uploaded
if uploaded_file and not st.session_state.file_uploaded:
    st.session_state.processed = False
    st.session_state.file_uploaded = True
    st.session_state.df = None
    st.session_state.output_data = None
    st.session_state.output_data_csv = None
    st.session_state.clean_output_data = None
    st.session_state.clean_output_data_csv = None
    st.session_state.clean_df = None
    st.session_state.clean_count = 0
    st.session_state.duplicate_count = 0
    st.session_state.duplicate_count_before = 0
    st.session_state.invalid_count = 0
    st.session_state.valid_count = 0
    st.session_state.column = None
    st.session_state.sheet_names = None
    st.session_state.selected_sheet = None
    st.session_state.original_file = uploaded_file

if uploaded_file:
    try:
        # Check if it's an Excel file and get sheet names
        if uploaded_file.name.endswith(".xlsx"):
            if st.session_state.sheet_names is None:
                # Read sheet names
                excel_file = pd.ExcelFile(uploaded_file)
                st.session_state.sheet_names = excel_file.sheet_names
                st.session_state.selected_sheet = excel_file.sheet_names[0] if excel_file.sheet_names else None
            
            # Sheet selection
            if st.session_state.sheet_names:
                st.markdown("### 📑 Sheet Selection")
                selected_sheet = st.selectbox(
                    "Select Sheet to Process",
                    st.session_state.sheet_names,
                    help="Choose which sheet contains your mobile number data",
                    key="sheet_selector"
                )
                
                # Check if sheet selection changed
                if st.session_state.selected_sheet != selected_sheet:
                    st.session_state.selected_sheet = selected_sheet
                    st.session_state.df = None  # Reset df to force reload
                    st.session_state.processed = False  # Reset processed state
            
            # Read the selected sheet only if df is None or sheet changed
            if st.session_state.df is None:
                with st.spinner(f"📂 Reading sheet '{st.session_state.selected_sheet}'..."):
                    df = pd.read_excel(uploaded_file, sheet_name=st.session_state.selected_sheet, dtype=str)
                    df = df.fillna("")
                    st.session_state.df = df
            else:
                df = st.session_state.df
        else:
            # CSV file
            if st.session_state.df is None:
                with st.spinner("📂 Reading CSV file..."):
                    df = pd.read_csv(uploaded_file, dtype=str, low_memory=False)
                    df = df.fillna("")
                    st.session_state.df = df
            else:
                df = st.session_state.df
        
        st.markdown("---")
        st.subheader("📊 File Preview")
        st.dataframe(df.head(10), use_container_width=True)
        st.caption(f"📌 Total rows: {len(df):,}")
        
        # Auto-detect phone number columns
        phone_keywords = ['phone', 'mobile', 'cell', 'number', 'contact', 'tel', 'mob']
        phone_columns = [col for col in df.columns if any(keyword in col.lower() for keyword in phone_keywords)]
        
        col1, col2 = st.columns([1, 1])
        
        with col1:
            column = st.selectbox(
                "📞 Select Mobile Number Column",
                df.columns,
                index=df.columns.get_loc(phone_columns[0]) if phone_columns else 0,
                help="Select the column containing the mobile numbers to validate",
                key="column_select"
            )
        
        # Validation options
        st.markdown("### ⚙️ Validation Settings")
        
        col1, col2 = st.columns(2)
        
        with col1:
            validation_type = st.radio(
                "📋 Validation Type:",
                ["Indian Mobile (10 digits, starts 6-9)", 
                 "International (with country code)"],
                help="Choose the validation rules for mobile numbers",
                key="validation_type"
            )
        
        with col2:
            show_detailed_reasons = st.checkbox(
                "📝 Show detailed invalid reasons", 
                value=True,
                help="Shows specific reason why a number is invalid",
                key="show_detailed"
            )
            
            keep_first = st.checkbox(
                "💾 Keep first occurrence of duplicates", 
                value=True,
                help="If checked, the first occurrence is kept as valid, rest marked as duplicate",
                key="keep_first"
            )
            
            remove_duplicates = st.checkbox(
                "🗑️ Remove duplicate records from output", 
                value=False,
                help="Keep only unique valid numbers (removes all duplicates)",
                key="remove_dups"
            )
        
        # Process button
        if st.button("🔄 Process File", type="primary", key="process_button"):
            st.session_state.processed = True
            st.session_state.output_data = None
            st.session_state.output_data_csv = None
            st.session_state.clean_output_data = None
            st.session_state.clean_output_data_csv = None
            
            # Create progress bar
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            with st.spinner("🔄 Processing mobile numbers..."):
                
                # Step 1: Clean values (5%)
                status_text.text("🧹 Cleaning numbers...")
                progress_bar.progress(5)
                
                # Clean values using vectorized operations
                df[column] = df[column].astype(str).str.strip()
                
                # Remove spaces, dashes, brackets, plus signs - vectorized
                df['cleaned_number'] = df[column].str.replace(r'[\s\-\(\)\+]', '', regex=True)
                
                # Step 2: Basic validation (15%)
                status_text.text("✅ Validating number formats...")
                progress_bar.progress(15)
                
                # Define patterns based on validation type
                if validation_type == "Indian Mobile (10 digits, starts 6-9)":
                    mobile_pattern = r"^[6-9]\d{9}$"
                    
                    # Vectorized operations for Indian numbers
                    cleaned_series = df['cleaned_number']
                    
                    # Check if number is exactly 10 digits
                    is_10_digits = cleaned_series.str.fullmatch(r'^\d{10}$', na=False)
                    
                    # Basic valid mask - using numpy for faster operations
                    valid_mask = cleaned_series.str.fullmatch(mobile_pattern, na=False)
                    
                    # Landline indicators - ONLY apply to 10-digit numbers
                    starts_with_0 = cleaned_series.str.startswith('0', na=False)
                    starts_with_1_5 = cleaned_series.str.match(r'^[1-5]', na=False)
                    
                    # ONLY mark as landline if it's exactly 10 digits and starts with 0 or 1-5
                    # OR if it starts with 0 and has more than 10 digits (STD code)
                    landline_mask = pd.Series(False, index=df.index)
                    
                    # For 10-digit numbers: starts with 0 or 1-5 are landlines
                    landline_mask = landline_mask | (is_10_digits & (starts_with_0 | starts_with_1_5))
                    
                    # For numbers with STD codes (starts with 0 and length > 10)
                    starts_with_0_long = cleaned_series.str.startswith('0', na=False) & (cleaned_series.str.len() > 10)
                    landline_mask = landline_mask | starts_with_0_long
                    
                    # STD code detection - only for numbers that don't match mobile pattern
                    std_codes = ('011', '022', '033', '044', '020', '080', '040', '0120', '079', '0755', 
                                '0124', '0129', '0172', '0183', '0161', '0164', '0141', '0145')
                    
                    # Vectorized STD code check using a single combined pattern
                    std_pattern = '|'.join([f'^{code}' for code in std_codes])
                    std_code_mask = cleaned_series.str.match(std_pattern, na=False) & ~valid_mask
                    
                    # Update landline mask
                    landline_mask = landline_mask | std_code_mask
                    
                    # Update valid mask - must be valid mobile AND not landline
                    valid_mask = valid_mask & ~landline_mask
                    
                else:  # International
                    mobile_pattern = r"^\d{10,15}$"
                    valid_mask = df['cleaned_number'].str.fullmatch(mobile_pattern, na=False)
                    landline_mask = pd.Series(False, index=df.index)
                    std_code_mask = pd.Series(False, index=df.index)
                
                # Step 3: Find duplicates (35%)
                status_text.text("🔍 Checking for duplicates...")
                progress_bar.progress(35)
                
                # FIXED: Find duplicates among valid numbers
                duplicate_mask = pd.Series(False, index=df.index)
                
                if valid_mask.any():
                    # Get valid numbers with their indices
                    valid_indices = df[valid_mask].index
                    valid_numbers = df.loc[valid_mask, 'cleaned_number']
                    
                    # Count occurrences of each number
                    value_counts = valid_numbers.value_counts()
                    
                    # Find numbers that appear more than once
                    duplicated_values = value_counts[value_counts > 1].index
                    
                    if len(duplicated_values) > 0:
                        # Create a series to mark duplicates
                        if keep_first:
                            # Mark duplicates (keep first, mark rest)
                            # Use duplicated() on the valid numbers series
                            dup_series = valid_numbers.duplicated(keep='first')
                            # Only mark duplicates for numbers that appear more than once
                            dup_series = dup_series & valid_numbers.isin(duplicated_values)
                        else:
                            # Mark all occurrences as duplicate
                            dup_series = valid_numbers.isin(duplicated_values)
                        
                        # Assign back to duplicate_mask using proper index alignment
                        duplicate_mask.loc[valid_indices] = dup_series.values
                
                # Step 4: Create remarks (60%)
                status_text.text("📝 Generating remarks...")
                progress_bar.progress(60)
                
                # Create remarks - vectorized
                if show_detailed_reasons:
                    remarks = pd.Series("", index=df.index)
                    
                    # Empty/blank checks
                    empty_mask = (df[column] == "") | (df[column].str.strip() == "")
                    remarks[empty_mask] = "Empty/Blank"
                    
                    # Detailed invalid reasons for Indian numbers
                    if validation_type == "Indian Mobile (10 digits, starts 6-9)":
                        cleaned = df['cleaned_number']
                        cleaned_len = cleaned.str.len()
                        
                        # First, check for landline indicators (only for valid length numbers)
                        # 10-digit numbers starting with 0 or 1-5
                        is_10_digits = cleaned.str.fullmatch(r'^\d{10}$', na=False)
                        
                        # Landline detection
                        landline_10_digit = is_10_digits & (cleaned.str.startswith('0', na=False) | cleaned.str.match(r'^[1-5]', na=False))
                        remarks[landline_10_digit] = "Landline (starts with 0 or 1-5)"
                        
                        # Numbers starting with 0 and longer than 10 digits (STD codes)
                        landline_std = cleaned.str.startswith('0', na=False) & (cleaned.str.len() > 10)
                        remarks[landline_std] = "Landline (STD code detected)"
                        
                        # STD code detection for other landlines
                        std_pattern = '|'.join([f'^{code}' for code in ('011', '022', '033', '044', '020', '080', '040', '0120', '079', '0755', '0124', '0129', '0172', '0183', '0161', '0164', '0141', '0145')])
                        std_code_landline = cleaned.str.match(std_pattern, na=False) & ~valid_mask & (remarks == "")
                        remarks[std_code_landline] = "Landline (STD code detected)"
                        
                        # Length issues
                        remarks[cleaned.str.len() < 10] = "Too Short"
                        remarks[cleaned.str.len() > 15] = "Too Long"
                        
                        # Numbers with letters or special characters
                        has_letters = cleaned.str.contains(r'[A-Za-z]', na=False)
                        remarks[has_letters & (remarks == "")] = "Invalid Format"
                    
                    # Invalid format (catch all)
                    invalid_format_mask = ~valid_mask & (remarks == "")
                    remarks[invalid_format_mask] = "Invalid Format"
                    
                    # Duplicates
                    remarks[duplicate_mask] = "Duplicate"
                else:
                    remarks = pd.Series("", index=df.index)
                    remarks[(df[column] == "") | (df[column].str.strip() == "")] = "Invalid Number"
                    remarks[~valid_mask] = "Invalid Number"
                    remarks[duplicate_mask] = "Duplicate"
                
                df["Remarks"] = remarks
                
                # Step 5: Handle duplicate removal (80%)
                status_text.text("🗑️ Processing duplicate removal...")
                progress_bar.progress(80)
                
                # Track duplicate counts before removal
                duplicate_count_before = duplicate_mask.sum()
                
                # Handle duplicate removal
                if remove_duplicates:
                    df = df[~duplicate_mask].copy()
                    # Recalculate remarks for remaining records
                    df["Remarks"] = df["Remarks"].where(df["Remarks"] != "Duplicate", "")
                
                # Step 6: Calculate statistics (90%)
                status_text.text("📊 Calculating statistics...")
                progress_bar.progress(90)
                
                # Counts - optimized
                duplicate_count = (df["Remarks"] == "Duplicate").sum()
                invalid_count = ((df["Remarks"] != "") & (df["Remarks"] != "Duplicate")).sum()
                valid_count = len(df) - duplicate_count - invalid_count
                
                # Create clean dataset (valid numbers only, no duplicates)
                clean_df = df[(df["Remarks"] == "")].copy()
                # Remove the cleaned_number column if it exists
                if 'cleaned_number' in clean_df.columns:
                    clean_df = clean_df.drop(columns=['cleaned_number'])
                clean_count = len(clean_df)
                
                # Store processed data in session state
                st.session_state.df = df
                st.session_state.clean_df = clean_df
                st.session_state.duplicate_count = duplicate_count
                st.session_state.duplicate_count_before = duplicate_count_before
                st.session_state.invalid_count = invalid_count
                st.session_state.valid_count = valid_count
                st.session_state.clean_count = clean_count
                st.session_state.column = column
                
                # Complete
                status_text.text("✅ Done!")
                progress_bar.progress(100)
                
                # Display results
                st.success("✅ Processing completed successfully!")
                
                # Summary metrics with custom styling
                st.markdown("### 📊 Summary Statistics")
                
                col1, col2, col3, col4, col5 = st.columns(5)
                
                with col1:
                    st.markdown(f"""
                    <div class="metric-container">
                        <div class="metric-icon">📊</div>
                        <div class="metric-value">{len(df):,}</div>
                        <div class="metric-label">Total Records</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col2:
                    st.markdown(f"""
                    <div class="metric-container">
                        <div class="metric-icon">✅</div>
                        <div class="metric-value" style="color: #28a745;">{valid_count:,}</div>
                        <div class="metric-label">Valid Numbers <span class="badge badge-success">{valid_count/len(df)*100:.1f}%</span></div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col3:
                    st.markdown(f"""
                    <div class="metric-container">
                        <div class="metric-icon">🔄</div>
                        <div class="metric-value" style="color: #dc3545;">{duplicate_count:,}</div>
                        <div class="metric-label">Duplicates <span class="badge badge-danger">{duplicate_count/len(df)*100:.1f}%</span></div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col4:
                    st.markdown(f"""
                    <div class="metric-container">
                        <div class="metric-icon">❌</div>
                        <div class="metric-value" style="color: #ffc107;">{invalid_count:,}</div>
                        <div class="metric-label">Invalid Numbers <span class="badge badge-warning">{invalid_count/len(df)*100:.1f}%</span></div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col5:
                    st.markdown(f"""
                    <div class="metric-container">
                        <div class="metric-icon">✨</div>
                        <div class="metric-value" style="color: #17a2b8;">{clean_count:,}</div>
                        <div class="metric-label">Clean Records <span class="badge badge-info">{clean_count/len(df)*100:.1f}%</span></div>
                    </div>
                    """, unsafe_allow_html=True)
                
                # Show duplicate details
                with st.expander("🔄 View Duplicate Details", expanded=False):
                    duplicate_sample = df[df["Remarks"] == "Duplicate"]
                    if not duplicate_sample.empty:
                        st.write(f"**Found {duplicate_count:,} duplicate records** (all occurrences after the first)")
                        
                        # Show which numbers have duplicates
                        dup_numbers = duplicate_sample['cleaned_number'].unique()
                        st.write(f"**Numbers with duplicates:** {len(dup_numbers)} unique numbers")
                        
                        # Show detailed view with row numbers
                        dup_details = df[df.duplicated(subset=['cleaned_number'], keep=False)].copy()
                        if not dup_details.empty:
                            # Add occurrence count
                            dup_details['Occurrence'] = dup_details.groupby('cleaned_number').cumcount() + 1
                            dup_details['Total_Occurrences'] = dup_details.groupby('cleaned_number')['cleaned_number'].transform('count')
                            dup_details['Status'] = dup_details.apply(
                                lambda x: '✅ Keep (First)' if x['Occurrence'] == 1 and x['Remarks'] != 'Duplicate' else '🗑️ Remove (Duplicate)',
                                axis=1
                            )
                            
                            st.dataframe(dup_details[[column, 'cleaned_number', 'Occurrence', 'Total_Occurrences', 'Status', 'Remarks']], use_container_width=True)
                            
                            # Show summary of duplicates
                            dup_summary = dup_details.groupby('cleaned_number').agg({
                                'Occurrence': 'count'
                            }).reset_index()
                            dup_summary.columns = ['Mobile Number', 'Total Occurrences']
                            st.write("**Summary of duplicates:**")
                            st.dataframe(dup_summary, use_container_width=True)
                    else:
                        st.info("✅ No duplicate numbers found!")
                
                # Show invalid sample
                with st.expander("📋 View Invalid Numbers Sample", expanded=False):
                    invalid_sample = df[df["Remarks"].isin(["Invalid Number", "Invalid Format", "Empty/Blank", 
                                                            "Landline (starts with 0 or 1-5)", "Landline (STD code detected)",
                                                            "Too Short", "Too Long"])]
                    if not invalid_sample.empty:
                        st.dataframe(invalid_sample[[column, "Remarks"]].head(20), use_container_width=True)
                        st.caption(f"Showing first 20 of {invalid_count:,} invalid records")
                    else:
                        st.info("🎉 No invalid numbers found! All numbers are valid.")
                
                # Show clean data sample
                with st.expander("✨ View Clean Records Sample (Valid & Unique)", expanded=False):
                    if not clean_df.empty:
                        st.dataframe(clean_df[[column]].head(20), use_container_width=True)
                        st.caption(f"Showing first 20 of {clean_count:,} clean records")
                        st.success(f"📊 {clean_count:,} clean records ready for download!")
                    else:
                        st.warning("No clean records found!")
                
                # Prepare Excel outputs (95%)
                status_text.text("📁 Preparing files for download...")
                progress_bar.progress(95)
                
                # Prepare full processed file (Excel) - ALWAYS create even if empty
                output = BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="Result")
                    
                    ws = writer.sheets["Result"]
                    
                    # Define colors
                    duplicate_fill = PatternFill(
                        start_color="FFC7CE",
                        end_color="FFC7CE",
                        fill_type="solid"
                    )
                    
                    invalid_fill = PatternFill(
                        start_color="FFF2CC",
                        end_color="FFF2CC",
                        fill_type="solid"
                    )
                    
                    landline_fill = PatternFill(
                        start_color="FFE4B5",
                        end_color="FFE4B5",
                        fill_type="solid"
                    )
                    
                    valid_fill = PatternFill(
                        start_color="C6EFCE",
                        end_color="C6EFCE",
                        fill_type="solid"
                    )
                    
                    first_occurrence_fill = PatternFill(
                        start_color="D4E6F1",
                        end_color="D4E6F1",
                        fill_type="solid"
                    )
                    
                    # Find column indices
                    number_col = list(df.columns).index(column) + 1
                    remarks_col = list(df.columns).index("Remarks") + 1
                    
                    # Optimized color coding using numpy array
                    remarks_array = df['Remarks'].values
                    cleaned_array = df['cleaned_number'].values
                    
                    # Get duplicate numbers set for faster lookup
                    duplicate_numbers = set(df[df['Remarks'] == 'Duplicate']['cleaned_number'].values)
                    
                    # Color code rows - using batch processing
                    for idx, row_num in enumerate(range(2, len(df) + 2)):
                        remark = remarks_array[idx]
                        cleaned = cleaned_array[idx]
                        
                        if remark == "Duplicate":
                            ws.cell(row=row_num, column=number_col).fill = duplicate_fill
                            ws.cell(row=row_num, column=remarks_col).fill = duplicate_fill
                        elif "Landline" in str(remark):
                            ws.cell(row=row_num, column=number_col).fill = landline_fill
                            ws.cell(row=row_num, column=remarks_col).fill = landline_fill
                        elif remark in ["Invalid Number", "Invalid Format", "Empty/Blank", "Too Short", "Too Long"]:
                            ws.cell(row=row_num, column=number_col).fill = invalid_fill
                            ws.cell(row=row_num, column=remarks_col).fill = invalid_fill
                        elif remark == "":
                            # Check if this number has duplicates
                            if cleaned in duplicate_numbers:
                                ws.cell(row=row_num, column=number_col).fill = first_occurrence_fill
                                ws.cell(row=row_num, column=remarks_col).fill = first_occurrence_fill
                            else:
                                ws.cell(row=row_num, column=number_col).fill = valid_fill
                                ws.cell(row=row_num, column=remarks_col).fill = valid_fill
                    
                    # Auto-adjust column widths (optimized)
                    for col in ws.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[col_letter].width = adjusted_width
                
                output.seek(0)
                st.session_state.output_data = output.getvalue()
                
                # Prepare full processed file (CSV) - ALWAYS create even if empty
                csv_buffer = BytesIO()
                df.to_csv(csv_buffer, index=False, encoding='utf-8')
                csv_buffer.seek(0)
                st.session_state.output_data_csv = csv_buffer.getvalue()
                
                # Prepare clean file (Excel) - only if clean records exist
                if not clean_df.empty:
                    clean_output = BytesIO()
                    with pd.ExcelWriter(clean_output, engine="openpyxl") as writer:
                        clean_df.to_excel(writer, index=False, sheet_name="Clean Numbers")
                        
                        ws = writer.sheets["Clean Numbers"]
                        
                        # Auto-adjust column widths
                        for col in ws.columns:
                            max_length = 0
                            col_letter = col[0].column_letter
                            for cell in col:
                                try:
                                    if cell.value:
                                        max_length = max(max_length, len(str(cell.value)))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[col_letter].width = adjusted_width
                    
                    clean_output.seek(0)
                    st.session_state.clean_output_data = clean_output.getvalue()
                    
                    # Prepare clean file (CSV)
                    csv_buffer = BytesIO()
                    clean_df.to_csv(csv_buffer, index=False, encoding='utf-8')
                    csv_buffer.seek(0)
                    st.session_state.clean_output_data_csv = csv_buffer.getvalue()
                else:
                    # Set to None if no clean records
                    st.session_state.clean_output_data = None
                    st.session_state.clean_output_data_csv = None
                
                # Clear progress indicators
                progress_bar.empty()
                status_text.empty()
                
                # Download buttons section
                st.markdown("---")
                st.markdown("### 📥 Download Options")
                
                # Row 1: Full Processed Files
                st.markdown("#### 📊 Full Processed File (With Remarks)")
                col1, col2 = st.columns(2)
                
                with col1:
                    st.download_button(
                        label="📥 Download Excel (.xlsx)",
                        data=st.session_state.output_data,
                        file_name="Processed_Mobile_Numbers_With_Remarks.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="download_excel_full"
                    )
                
                with col2:
                    st.download_button(
                        label="📥 Download CSV (.csv)",
                        data=st.session_state.output_data_csv,
                        file_name="Processed_Mobile_Numbers_With_Remarks.csv",
                        mime="text/csv",
                        use_container_width=True,
                        key="download_csv_full"
                    )
                
                # Row 2: Clean Numbers Files
                if clean_count > 0 and st.session_state.clean_output_data is not None:
                    st.markdown("#### ✨ Clean Numbers Only (Valid & Unique)")
                    col3, col4 = st.columns(2)
                    
                    with col3:
                        st.download_button(
                            label="📥 Download Excel (.xlsx)",
                            data=st.session_state.clean_output_data,
                            file_name="Clean_Valid_Unique_Numbers.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key="download_excel_clean"
                        )
                    
                    with col4:
                        st.download_button(
                            label="📥 Download CSV (.csv)",
                            data=st.session_state.clean_output_data_csv,
                            file_name="Clean_Valid_Unique_Numbers.csv",
                            mime="text/csv",
                            use_container_width=True,
                            key="download_csv_clean"
                        )
                else:
                    st.info("No clean records available for download")
        
        # If already processed, show download buttons again without reprocessing
        elif st.session_state.processed and st.session_state.output_data is not None:
            st.success("✅ File already processed! Download your files below.")
            
            # Show metrics from session state
            if hasattr(st.session_state, 'valid_count'):
                st.markdown("### 📊 Summary Statistics")
                col1, col2, col3, col4, col5 = st.columns(5)
                
                with col1:
                    st.markdown(f"""
                    <div class="metric-container">
                        <div class="metric-icon">📊</div>
                        <div class="metric-value">{len(st.session_state.df):,}</div>
                        <div class="metric-label">Total Records</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col2:
                    st.markdown(f"""
                    <div class="metric-container">
                        <div class="metric-icon">✅</div>
                        <div class="metric-value" style="color: #28a745;">{st.session_state.valid_count:,}</div>
                        <div class="metric-label">Valid Numbers</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col3:
                    st.markdown(f"""
                    <div class="metric-container">
                        <div class="metric-icon">🔄</div>
                        <div class="metric-value" style="color: #dc3545;">{st.session_state.duplicate_count:,}</div>
                        <div class="metric-label">Duplicates</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col4:
                    st.markdown(f"""
                    <div class="metric-container">
                        <div class="metric-icon">❌</div>
                        <div class="metric-value" style="color: #ffc107;">{st.session_state.invalid_count:,}</div>
                        <div class="metric-label">Invalid Numbers</div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col5:
                    st.markdown(f"""
                    <div class="metric-container">
                        <div class="metric-icon">✨</div>
                        <div class="metric-value" style="color: #17a2b8;">{st.session_state.clean_count:,}</div>
                        <div class="metric-label">Clean Records</div>
                    </div>
                    """, unsafe_allow_html=True)
            
            st.markdown("---")
            st.markdown("### 📥 Download Options")
            
            # Row 1: Full Processed Files
            st.markdown("#### 📊 Full Processed File (With Remarks)")
            col1, col2 = st.columns(2)
            
            with col1:
                st.download_button(
                    label="📥 Download Excel (.xlsx)",
                    data=st.session_state.output_data,
                    file_name="Processed_Mobile_Numbers_With_Remarks.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="download_excel_full_again"
                )
            
            with col2:
                st.download_button(
                    label="📥 Download CSV (.csv)",
                    data=st.session_state.output_data_csv,
                    file_name="Processed_Mobile_Numbers_With_Remarks.csv",
                    mime="text/csv",
                    use_container_width=True,
                    key="download_csv_full_again"
                )
            
            # Row 2: Clean Numbers Files
            if st.session_state.clean_count > 0 and st.session_state.clean_output_data is not None:
                st.markdown("#### ✨ Clean Numbers Only (Valid & Unique)")
                col3, col4 = st.columns(2)
                
                with col3:
                    st.download_button(
                        label="📥 Download Excel (.xlsx)",
                        data=st.session_state.clean_output_data,
                        file_name="Clean_Valid_Unique_Numbers.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="download_excel_clean_again"
                    )
                
                with col4:
                    st.download_button(
                        label="📥 Download CSV (.csv)",
                        data=st.session_state.clean_output_data_csv,
                        file_name="Clean_Valid_Unique_Numbers.csv",
                        mime="text/csv",
                        use_container_width=True,
                        key="download_csv_clean_again"
                    )
            else:
                st.info("No clean records available for download")
                
    except Exception as e:
        st.error(f"❌ Error processing file: {str(e)}")
        st.info("Please check that your file is properly formatted and contains valid data.")
        # Reset session state on error
        st.session_state.processed = False
        st.session_state.df = None
        st.session_state.output_data = None
        st.session_state.output_data_csv = None
        st.session_state.clean_output_data = None
        st.session_state.clean_output_data_csv = None
        st.session_state.clean_df = None
        st.session_state.clean_count = 0
        st.session_state.duplicate_count = 0
        st.session_state.duplicate_count_before = 0
        st.session_state.invalid_count = 0
        st.session_state.valid_count = 0
        st.session_state.column = None

else:
    # Display empty state
    st.info("👆 Please upload an Excel or CSV file to get started")
    # Reset session state when no file
    st.session_state.processed = False
    st.session_state.file_uploaded = False
    st.session_state.df = None
    st.session_state.output_data = None
    st.session_state.output_data_csv = None
    st.session_state.clean_output_data = None
    st.session_state.clean_output_data_csv = None
    st.session_state.clean_df = None
    st.session_state.clean_count = 0
    st.session_state.duplicate_count = 0
    st.session_state.duplicate_count_before = 0
    st.session_state.invalid_count = 0
    st.session_state.valid_count = 0
    st.session_state.column = None
    st.session_state.sheet_names = None
    st.session_state.selected_sheet = None
    
    # Custom empty state with styling
    st.markdown("""
    <div style="text-align: center; padding: 3rem 0;">
        <div style="font-size: 4rem; margin-bottom: 1rem;">📂</div>
        <h3 style="color: #2c3e50;">Ready to Process Your Mobile Numbers</h3>
        <p style="color: #7f8c8d; max-width: 600px; margin: 1rem auto;">
            Upload your Excel or CSV file containing mobile numbers to get started.
            The tool will automatically detect duplicates, invalid numbers, and more.
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    with st.expander("📖 How to use this tool", expanded=False):
        st.markdown("""
        ### Step-by-step guide:
        1. **Upload** your Excel or CSV file containing mobile numbers
        2. **For Excel files**: Select the sheet containing your data
        3. **Select** the column that contains the mobile numbers
        4. **Choose** validation settings (Indian or International)
        5. **Click** the Process File button
        6. **Download** your files:
           - **Full Processed File**: Contains all records with color-coded remarks
           - **Clean Numbers Only**: Only valid, unique mobile numbers (perfect for use)
           - Available in both **Excel (.xlsx)** and **CSV (.csv)** formats
        
        ### Features:
        - ✅ Validates 10-digit Indian mobile numbers (starting with 6-9)
        - ❌ Identifies landline numbers (starting with 0 or 1-5)
        - 🔄 Detects duplicate numbers (keeps first, marks rest)
        - 🎨 Color-coded Excel output (Green=Valid, Red=Duplicate, Orange=Landline, Yellow=Invalid)
        - ✨ Download only clean, valid, unique numbers
        - 📊 Both Excel and CSV formats available
        - 📋 Detailed statistics and breakdowns
        - 🚀 Fast processing for large files
        - 📈 Progress bar shows processing status
        - 📑 Support for multi-sheet Excel files
        """)

st.markdown("---")
st.markdown("""
<div class="footer">
    Built with ❤️ using Streamlit | For any issues, please check your file format
</div>
""", unsafe_allow_html=True)